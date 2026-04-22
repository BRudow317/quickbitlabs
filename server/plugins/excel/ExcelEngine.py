"""
Excel engine for the Excel plugin.

File model: one workbook (.xlsx) = one Catalog; one worksheet = one Entity.

Reading:  polars.read_excel with openpyxl engine — handles type inference and
          returns a polars DataFrame which we convert to a PyArrow table.

Writing:  openpyxl directly — gives full workbook control (add/replace individual
          sheets without disturbing others) which polars write_excel cannot guarantee
          in append mode.

Schema inference note: Excel has no embedded schema metadata (unlike Parquet/Feather).
Both read_schema and read must load the full worksheet to infer types correctly.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import polars as pl
import pyarrow as pa


class ExcelEngine:

    @staticmethod
    def list_sheets(path: Path) -> list[str]:
        """Return worksheet names using openpyxl read-only mode (no cell data loaded)."""
        wb = openpyxl.load_workbook(str(path), read_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    @staticmethod
    def read_schema(path: Path, sheet_name: str) -> pa.Schema:
        """Infer schema by reading the full worksheet via polars."""
        return ExcelEngine.read(path, sheet_name).schema

    @staticmethod
    def read(path: Path, sheet_name: str) -> pa.RecordBatchReader:
        """Read a worksheet into a RecordBatchReader via polars type inference."""
        df = pl.read_excel(
            path,
            sheet_name=sheet_name,
            engine="openpyxl",
            infer_schema_length=None,
        )
        table = df.to_arrow()
        return pa.RecordBatchReader.from_batches(table.schema, table.to_batches())

    @staticmethod
    def write_sheet(path: Path, sheet_name: str, stream: pa.RecordBatchReader) -> None:
        """
        Write/replace a worksheet in the workbook using openpyxl.
        Creates the file if it does not exist; preserves all other sheets.
        """
        table = stream.read_all()

        if path.exists():
            wb = openpyxl.load_workbook(str(path))
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
        else:
            wb = openpyxl.Workbook()
            # Remove the blank "Sheet" openpyxl always creates on a new workbook
            for default in list(wb.sheetnames):
                del wb[default]

        ws = wb.create_sheet(sheet_name)

        # Header row
        ws.append([f.name for f in table.schema])

        # Data rows — convert each column to Python list once, then zip across columns
        col_lists = [table.column(i).to_pylist() for i in range(table.num_columns)]
        for row in zip(*col_lists):
            ws.append(list(row))

        wb.save(str(path))

    @staticmethod
    def delete_sheet(path: Path, sheet_name: str) -> None:
        """Remove a worksheet from the workbook. No-op if the file or sheet does not exist."""
        if not path.exists():
            return
        wb = openpyxl.load_workbook(str(path))
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            wb.save(str(path))
